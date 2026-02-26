from __future__ import annotations

import argparse
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# -------------------------
# Excel column indices (1-based)
# -------------------------
COL_GENE = 5            # E
COL_VARIANT = 6         # F (Variant_curation)
COL_MODEL = 7           # G (FindingsModel_curation)
COL_ARGS = 9            # I


# =========================
# Errors
# =========================

class MappingError(ValueError):
    """Raised when a row cannot be deterministically mapped."""


# =========================
# Low-level split utilities
# =========================

def normalize_text(v: object) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def split_top_level(s: str, sep: str) -> List[str]:
    """
    Split by `sep` only when not inside parentheses.
    Example: "A,(B,C),D" split by ',' -> ["A", "(B,C)", "D"]
    """
    s = s or ""
    parts: List[str] = []
    buf: List[str] = []
    depth = 0

    for ch in s:
        if ch == "(":
            depth += 1
            buf.append(ch)
        elif ch == ")":
            depth = max(0, depth - 1)
            buf.append(ch)
        elif ch == sep and depth == 0:
            part = "".join(buf).strip()
            if part:
                parts.append(part)
            buf = []
        else:
            buf.append(ch)

    tail = "".join(buf).strip()
    if tail:
        parts.append(tail)
    return parts


def split_commas(s: str) -> List[str]:
    return split_top_level(s, ",")


def split_ands(s: str) -> List[str]:
    return split_top_level(s, "&")


# =========================
# FindingsModel expression AST
# =========================

@dataclass(frozen=True)
class Expr:
    pass


@dataclass(frozen=True)
class Atom(Expr):
    token: str  # e.g. "SmallVariant" or "GainDeletion.type.SOMATIC_GAIN"


@dataclass(frozen=True)
class Not(Expr):
    child: Expr


@dataclass(frozen=True)
class And(Expr):
    left: Expr
    right: Expr


@dataclass(frozen=True)
class Or(Expr):
    left: Expr
    right: Expr


_TOKEN_RE = re.compile(
    r"""
    \s*
    (
        NOT\b
      | \(
      | \)
      | \|
      | &
      | [A-Za-z][A-Za-z0-9_.]*   # class token with optional .type.QUAL
    )
    """,
    re.VERBOSE | re.IGNORECASE,
)


def tokenize_model_expr(s: str) -> List[str]:
    if not s:
        raise MappingError("Empty FindingsModel_curation expression cannot be tokenized.")
    tokens: List[str] = []
    pos = 0
    while pos < len(s):
        m = _TOKEN_RE.match(s, pos)
        if not m:
            raise MappingError(f"Cannot tokenize FindingsModel_curation near: {s[pos:pos+40]!r}")
        tok = m.group(1)
        pos = m.end()
        tokens.append("NOT" if tok.upper() == "NOT" else tok)
    return tokens


def _precedence(op: str) -> int:
    # NOT > & > |
    return {"NOT": 3, "&": 2, "|": 1}.get(op, 0)


def parse_model_expr(s: str) -> Expr:
    """
    Parse boolean expression over findings tokens with:
      - NOT(...)
      - &
      - |
      - parentheses
    """
    tokens = tokenize_model_expr(s)
    ops: List[str] = []
    vals: List[Expr] = []

    def apply_op(op: str) -> None:
        if op == "NOT":
            if not vals:
                raise MappingError("NOT has no operand")
            vals.append(Not(vals.pop()))
            return
        if len(vals) < 2:
            raise MappingError(f"Operator {op} missing operand(s)")
        b = vals.pop()
        a = vals.pop()
        vals.append(And(a, b) if op == "&" else Or(a, b))

    for t in tokens:
        if t == "(":
            ops.append(t)
        elif t == ")":
            while ops and ops[-1] != "(":
                apply_op(ops.pop())
            if not ops:
                raise MappingError("Mismatched parentheses")
            ops.pop()  # "("
        elif t in ("NOT", "&", "|"):
            if t == "NOT":
                ops.append(t)
            else:
                while ops and ops[-1] != "(" and _precedence(ops[-1]) >= _precedence(t):
                    apply_op(ops.pop())
                ops.append(t)
        else:
            vals.append(Atom(t))
            # apply any immediately pending NOT
            while ops and ops[-1] == "NOT":
                apply_op(ops.pop())

    while ops:
        op = ops.pop()
        if op in ("(", ")"):
            raise MappingError("Mismatched parentheses")
        apply_op(op)

    if len(vals) != 1:
        raise MappingError(f"Invalid expression: {s!r}")
    return vals[0]


def push_down_not(expr: Expr) -> Expr:
    """
    Convert to negation normal form (NNF):
    - push NOT down to atoms using De Morgan
    - eliminate double negations
    """
    if isinstance(expr, Atom):
        return expr
    if isinstance(expr, Not):
        c = expr.child
        if isinstance(c, Atom):
            return expr
        if isinstance(c, Not):
            return push_down_not(c.child)
        if isinstance(c, And):
            return Or(push_down_not(Not(c.left)), push_down_not(Not(c.right)))
        if isinstance(c, Or):
            return And(push_down_not(Not(c.left)), push_down_not(Not(c.right)))
        return Not(push_down_not(c))
    if isinstance(expr, And):
        return And(push_down_not(expr.left), push_down_not(expr.right))
    if isinstance(expr, Or):
        return Or(push_down_not(expr.left), push_down_not(expr.right))
    raise TypeError(expr)


# =========================
# Normalization (FindingsModel token)
# =========================

_CLASS_MAP = {
    "virus": "Virus",
    "arm": "Arm",
    "fusion": "Fusion",
    "disruption": "Disruption",
    "gaindeletion": "GainDeletion",
    "smallvariant": "SmallVariant",
}


def normalize_model_token(token: str) -> str:
    """
    Normalize casing for class names and qualifiers.
    Examples:
      gaindeletion.type.somatic_gain -> GainDeletion.type.SOMATIC_GAIN
      Smallvariant -> SmallVariant
      GainDeletion.Type.Somatic_Del -> GainDeletion.type.SOMATIC_DEL
    """
    t = token.strip()
    parts = t.split(".")
    if not parts:
        return t

    cls = _CLASS_MAP.get(parts[0].lower(), parts[0])

    if len(parts) == 1:
        return cls

    if parts[1].lower() == "type":
        qual = parts[2].upper() if len(parts) >= 3 else ""
        return f"{cls}.type.{qual}".rstrip(".")

    return ".".join([cls] + parts[1:])


# =========================
# Atomic mapping helpers
# =========================

HGVS_PROTEIN_RE = re.compile(r"\bp\.[A-Za-z0-9_?*]+")
EXON_RE = re.compile(r"\bexon\s+(\d+)\b", re.IGNORECASE)

_EFFECT_KEYWORDS: List[Tuple[str, str]] = [
    ("missense", "MISSENSE"),
    ("synonymous", "SYNONYMOUS"),
    ("stop gained", "STOP_GAINED"),
    ("nonsense", "STOP_GAINED"),
    ("stop lost", "STOP_LOST"),
    ("start lost", "START_LOST"),
]


def format_args_kv(pairs: List[Tuple[str, str]]) -> str:
    return " & ".join(f"{k}={v}" for k, v in pairs)


def infer_arm_type(variant_curation: str) -> str:
    v = variant_curation.lower()
    return "ARM_GAIN" if any(x in v for x in ("amplification", "copy number gain", "trisomy")) else "ARM_LOSS"


def parse_arm_locus(locus: str) -> Dict[str, str]:
    """
    Parse:
      1p, 16q, 1q21, 1q21.3 (sub-band ignored)
    Returns keys among: chromosome, arm, region, band
    """
    s = locus.strip()
    s = re.sub(r"(\d+[pq]\d+)\.\d+$", r"\1", s, flags=re.IGNORECASE)  # ignore sub-band
    m = re.match(r"^(\d+)\s*([pq])?\s*([0-9]+)?$", s, flags=re.IGNORECASE)
    if not m:
        raise MappingError(f"Unparsable Arm locus: {locus!r}")

    chrom = m.group(1)
    arm = m.group(2).lower() if m.group(2) else None
    digits = m.group(3)

    out: Dict[str, str] = {"chromosome": chrom}
    if arm:
        out["arm"] = arm
    if digits and len(digits) >= 2:
        out["region"] = digits[0]
        out["band"] = digits[1]
    return out


def infer_gaindel_type_from_variant_curation(variant_curation: str) -> Optional[str]:
    """
    GainDeletion type inference from Variant_curation (col F).
    - deletion/loss -> SOMATIC_DEL
    - amplification/copy number gain/cn gain -> SOMATIC_GAIN
    """
    v = (variant_curation or "").lower()
    # Order matters: DEL overrides gain if both appear.
    if "deletion" in v or "loss" in v:
        return "SOMATIC_DEL"
    if "amplification" in v or "copy number gain" in v or "cn gain" in v:
        return "SOMATIC_GAIN"
    return None


# =========================
# Atomic mapping
# =========================

def map_atom_to_args(atom_token: str, gene: str, variant_curation: str) -> str:
    """
    Returns 'Class[...]' (without leading '!').

    GainDeletion precedence:
      - forced GainDeletion.type.* in col G takes precedence (even if nested under NOT(...))
      - otherwise infer from Variant_curation (col F)
    """
    token = normalize_model_token(atom_token)

    forced_type: Optional[str] = None
    if token.startswith("GainDeletion.type."):
        forced_type = token.split(".")[-1].upper()
        cls = "GainDeletion"
    else:
        cls = token.split(".")[0]

    if cls == "Virus":
        return f"Virus[{format_args_kv([('name', gene)])}]"

    if cls == "Disruption":
        return f"Disruption[{format_args_kv([('gene', gene)])}]"

    if cls == "GainDeletion":
        pairs: List[Tuple[str, str]] = [("gene", gene)]
        typ = forced_type or infer_gaindel_type_from_variant_curation(variant_curation)
        if typ:
            pairs.append(("type", typ))
        return f"GainDeletion[{format_args_kv(pairs)}]"

    if cls == "SmallVariant":
        pairs: List[Tuple[str, str]] = [("gene", gene)]
        v = variant_curation
        v_lc = v.lower()

        m = HGVS_PROTEIN_RE.search(v)
        if m:
            pairs.append(("transcriptImpact.hgvsProteinImpact", m.group(0)))

        exon_num: Optional[str] = None
        m2 = EXON_RE.search(v)
        if m2:
            exon_num = m2.group(1)
            pairs.append(("transcriptImpact.affectedExon", exon_num))

        # effects
        if exon_num and "insertion" in v_lc:
            pairs.append(("transcriptImpact.effects", "INFRAME_INSERTION"))
        elif exon_num and "deletion" in v_lc:
            pairs.append(("transcriptImpact.effects", "INFRAME_DELETION"))
        else:
            for kw, enum in _EFFECT_KEYWORDS:
                if kw in v_lc:
                    pairs.append(("transcriptImpact.effects", enum))
                    break

        if "splice" in v_lc or "skipping" in v_lc:
            pairs.append(("transcriptImpact.inSpliceRegion", "True"))

        return f"SmallVariant[{format_args_kv(pairs)}]"

    if cls == "Fusion":
        if "_" in gene:
            g1, g2 = [x.strip() for x in gene.split("_", 1)]
            a = f"(geneStart={g1} & geneEnd={g2})"
            b = f"(geneStart={g2} & geneEnd={g1})"
            return f"Fusion[{a} | {b}]"
        return f"Fusion[geneStart={gene} | geneEnd={gene}]"

    if cls == "Arm":
        loci = [x.strip() for x in split_ands(gene)]
        arm_type = infer_arm_type(variant_curation)

        locus_chunks: List[str] = []
        for loc in loci:
            fields = parse_arm_locus(loc)
            kv: List[Tuple[str, str]] = [("chromosome", fields["chromosome"])]
            if "arm" in fields:
                kv.append(("arm", fields["arm"]))
            if "region" in fields:
                kv.append(("region", fields["region"]))
            if "band" in fields:
                kv.append(("band", fields["band"]))
            kv.append(("type", arm_type))

            locus_str = format_args_kv(kv)
            locus_chunks.append(locus_str)

        if len(locus_chunks) == 1:
            return f"Arm[{locus_chunks[0]}]"

        # Multiple loci: ( ... ) & ( ... )
        inner = " & ".join(f"({chunk})" for chunk in locus_chunks)
        return f"Arm[{inner}]"

    raise MappingError(f"Unknown findings class token: {atom_token!r} (normalized: {token!r})")


# =========================
# Render AST to output string
# =========================

def _needs_parens_under_and(expr: Expr) -> bool:
    return isinstance(expr, Or)


def _needs_parens_under_or(expr: Expr) -> bool:
    return isinstance(expr, And)


def expr_to_string(expr: Expr, gene: str, variant_curation: str) -> str:
    if isinstance(expr, Atom):
        return map_atom_to_args(expr.token, gene, variant_curation)

    if isinstance(expr, Not):
        if isinstance(expr.child, Atom):
            return "!" + map_atom_to_args(expr.child.token, gene, variant_curation)
        # Defensive fallback (shouldn't happen after push_down_not)
        return "!" + f"({expr_to_string(expr.child, gene, variant_curation)})"

    if isinstance(expr, And):
        left = expr_to_string(expr.left, gene, variant_curation)
        right = expr_to_string(expr.right, gene, variant_curation)
        if _needs_parens_under_and(expr.left):
            left = f"({left})"
        if _needs_parens_under_and(expr.right):
            right = f"({right})"
        return f"{left} & {right}"

    if isinstance(expr, Or):
        left = expr_to_string(expr.left, gene, variant_curation)
        right = expr_to_string(expr.right, gene, variant_curation)
        if _needs_parens_under_or(expr.left):
            left = f"({left})"
        if _needs_parens_under_or(expr.right):
            right = f"({right})"
        return f"{left} | {right}"

    raise TypeError(expr)


def _has_bool_ops(s: str) -> bool:
    return (" | " in s) or (" & " in s)


# =========================
# Broadcasting / row mapping
# =========================

def broadcast_align(genes: List[str], variants: List[str], models: List[str]) -> List[Tuple[str, str, str]]:
    g, v, m = len(genes), len(variants), len(models)
    sizes = (g, v, m)
    max_n = max(sizes)

    if g == v == m == 1:
        return [(genes[0], variants[0], models[0])]

    if g == v == m:
        return list(zip(genes, variants, models))

    # broadcast singleton(s) to max_n if compatible
    if any(n not in (1, max_n) for n in sizes):
        raise MappingError(f"Incompatible component counts (genes, variants, models) = {sizes}")

    def bcast(lst: List[str]) -> List[str]:
        return lst if len(lst) == max_n else [lst[0]] * max_n

    return list(zip(bcast(genes), bcast(variants), bcast(models)))


def map_row_to_args(gene_cell: object, variant_cell: object, model_cell: object) -> str:
    gene_s = normalize_text(gene_cell)
    variant_s = normalize_text(variant_cell)  # Variant_curation (F)
    model_s = normalize_text(model_cell)      # FindingsModel_curation (G)

    # If FindingsModel_curation is blank, ignore the row.
    if not model_s:
        return ""

    genes = split_commas(gene_s) if gene_s else [""]
    variants = split_commas(variant_s) if variant_s else [""]
    models = split_commas(model_s) if model_s else [""]

    triples = broadcast_align(genes, variants, models)

    outputs: List[str] = []
    for gene_i, var_i, model_i in triples:
        gene_i = normalize_text(gene_i)
        var_i = normalize_text(var_i)
        model_i = normalize_text(model_i)

        ast = push_down_not(parse_model_expr(model_i))
        rendered = expr_to_string(ast, gene_i, var_i)

        # If multiple comma-separated outputs and inner expression is compound, parenthesize.
        if len(triples) > 1 and _has_bool_ops(rendered):
            rendered = f"({rendered})"

        outputs.append(rendered)

    return ", ".join(outputs)


# =========================
# Excel I/O
# =========================

def get_sheet(wb, preferred_name: str = "Worksheet1") -> Worksheet:
    return wb[preferred_name] if preferred_name in wb.sheetnames else wb.active


def generate_mapping(
    xlsx_path: str,
    out_path: Optional[str] = None,
    sheet_name: str = "Worksheet1",
    start_row: int = 2,
    overwrite_col_i: bool = False,
    write_diff_sheet: bool = True,
) -> str:
    """
    Reads workbook and writes generated Args.

    If overwrite_col_i is True:
      - writes to column I (Args)
    Else:
      - writes to a new column at the end named "I_generated"

    Diff sheet:
      - If write_diff_sheet=True and there are mismatches or exceptions,
        writes a "MappingDiff" sheet.
    """
    wb = load_workbook(xlsx_path)
    ws = get_sheet(wb, sheet_name)

    # Determine output column
    if overwrite_col_i:
        target_col = COL_ARGS
    else:
        col = 1
        while ws.cell(row=1, column=col).value not in (None, ""):
            col += 1
        target_col = col
        ws.cell(row=1, column=target_col).value = "I_generated"

    diffs: List[Tuple[int, str, str, str]] = []  # (row, existing_I, generated, status)

    for r in range(start_row, ws.max_row + 1):
        gene = ws.cell(row=r, column=COL_GENE).value
        variant = ws.cell(row=r, column=COL_VARIANT).value
        model = ws.cell(row=r, column=COL_MODEL).value

        existing_i = ws.cell(row=r, column=COL_ARGS).value
        existing_i_s = normalize_text(existing_i)

        try:
            generated = map_row_to_args(gene, variant, model)
            ws.cell(row=r, column=target_col).value = generated

            # Only meaningful to diff if we are not overwriting I and I is populated.
            if (not overwrite_col_i) and existing_i_s and normalize_text(generated) != existing_i_s:
                diffs.append((r, existing_i_s, generated, "Mismatch"))

        except Exception as e:
            err = f"ERROR: {type(e).__name__}: {e}"
            ws.cell(row=r, column=target_col).value = err
            diffs.append((r, existing_i_s, err, "Exception"))

    if write_diff_sheet and diffs:
        if "MappingDiff" in wb.sheetnames:
            del wb["MappingDiff"]
        ds = wb.create_sheet("MappingDiff")
        ds.append(["Row", "Existing_I", "Generated", "Status"])
        for row_num, exist_i, gen, status in diffs:
            ds.append([row_num, exist_i, gen, status])

    if out_path is None:
        p = Path(xlsx_path)
        out_path = str(p.with_name(f"{p.stem}_mapped{p.suffix}"))

    wb.save(out_path)
    return out_path


# =========================
# CLI
# =========================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate Findings Args (col I) from cols E (Gene), F (Variant), G (FindingsModel)."
    )
    parser.add_argument(
        "--input_xlsx",
        type=Path,
        required=True,
        help="Path to input XLSX (expects cols E, F, G to be populated as per spec).",
    )
    parser.add_argument(
        "--output_dir",
        type=Path,
        required=True,
        help="Output directory to write the mapped XLSX.",
    )
    parser.add_argument(
        "--sheet_name",
        default="Worksheet1",
        help="Worksheet name to process (default: Worksheet1).",
    )
    parser.add_argument(
        "--overwrite_col_i",
        action="store_true",
        help="Overwrite column I (Args). If not set, writes to a new column 'I_generated'.",
    )
    parser.add_argument(
        "--no_diff_sheet",
        action="store_true",
        help="Disable writing the MappingDiff sheet.",
    )
    parser.add_argument(
        "--log_level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
        help="Logging level.",
    )

    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )

    args.output_dir.mkdir(parents=True, exist_ok=True)
    output_path = args.output_dir / f"{args.input_xlsx.stem}_mapped.xlsx"

    logger.info("Input XLSX: %s", args.input_xlsx)
    logger.info("Output XLSX: %s", output_path)
    logger.info("Sheet name: %s", args.sheet_name)
    logger.info("Overwrite col I: %s", args.overwrite_col_i)

    out = generate_mapping(
        xlsx_path=str(args.input_xlsx),
        out_path=str(output_path),
        sheet_name=args.sheet_name,
        overwrite_col_i=args.overwrite_col_i,
        write_diff_sheet=not args.no_diff_sheet,
    )

    logger.info("Done. Wrote: %s", out)


if __name__ == "__main__":
    main()
